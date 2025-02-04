import argparse
import glob
import os
import json
import logging
import pandas as pd
from datetime import datetime

def merge_excel_files(data_folder, merged_file, sheet_name=None, verbose=False):
    """
    Merge all Excel files in the given folder into a single DataFrame and save it.
    Reads a specific sheet (if provided) from each file; if sheet_name is None,
    it reads all sheets and concatenates them into a single DataFrame.
    """
    file_pattern = os.path.join(data_folder, "*.xlsx")
    files = glob.glob(file_pattern)
    
    if not files:
        raise FileNotFoundError(f"No Excel files found in folder: {data_folder}")
    
    all_data = []
    for file in files:
        if verbose:
            print(f"Reading file: {file}")
        # If sheet_name is None, read all sheets (returns a dict) and merge them.
        raw_data = pd.read_excel(file, engine="openpyxl", sheet_name=sheet_name)
        if isinstance(raw_data, dict):
            df = pd.concat(raw_data.values(), ignore_index=True)
        else:
            df = raw_data
        
        # Add a column indicating the source file.
        df["source_file"] = os.path.basename(file)
        all_data.append(df)
    
    merged_df = pd.concat(all_data, ignore_index=True)
    merged_df.to_excel(merged_file, index=False)
    if verbose:
        print(f"Merged file saved as '{merged_file}'\n")
    
    return merged_df, files

def process_data(df, args):
    """
    Clean and process the DataFrame using the provided arguments.
      - Standardizes column names.
      - Applies optional user-defined transformations from a JSON config.
      - Cleans data (drop or fill missing values).
      - Filters data by date and/or custom filters.
      - Calculates profit and (optionally) profit margin.
      - Creates a summary using the specified aggregation method.
    Returns the cleaned DataFrame and a summary DataFrame.
    """
    # Standardize column names to lowercase.
    df.columns = df.columns.str.lower()
    
    # Apply user-defined transformations from a JSON config if provided.
    if args.config:
        try:
            with open(args.config, 'r') as f:
                config = json.load(f)
            if "rename_columns" in config:
                df.rename(columns=config["rename_columns"], inplace=True)
            # Additional transformations can be added here.
            if args.verbose:
                print("Applied user-defined transformations from config.")
        except Exception as e:
            print(f"Error reading config file: {e}")
    
    # Data cleaning: either drop rows with missing values or fill them.
    if args.dropna:
        df.dropna(inplace=True)
        if args.verbose:
            print("Dropped rows with missing values.")
    elif args.fillna is not None:
        df.fillna(args.fillna, inplace=True)
        if args.verbose:
            print(f"Filled missing values with {args.fillna}.")
    
    # Date filtering (if start_date and end_date are provided).
    if args.start_date and args.end_date:
        date_col = args.date_column.lower() if args.date_column else "date"
        if date_col in df.columns:
            # Convert to datetime; non-parsable values become NaT.
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
            try:
                start_date = pd.to_datetime(args.start_date)
                end_date = pd.to_datetime(args.end_date)
            except Exception as e:
                raise ValueError("Incorrect date format. Please use YYYY-MM-DD.") from e
            df = df[(df[date_col] >= start_date) & (df[date_col] <= end_date)]
            if args.verbose:
                print(f"Filtered data between {start_date.date()} and {end_date.date()}.")
        else:
            print(f"Warning: Date column '{date_col}' not found; skipping date filtering.")
    
    # Custom filtering: Accept multiple filters in the format "column:value"
    if args.filter:
        for filt in args.filter:
            try:
                col, val = filt.split(":", 1)
                col = col.strip().lower()
                val = val.strip()
                df = df[df[col].astype(str) == val]
                if args.verbose:
                    print(f"Applied filter: {col} == {val}")
            except Exception as e:
                print(f"Error processing filter '{filt}': {e}")
    
    # Check for required columns to calculate profit.
    if "revenue" not in df.columns or "cost" not in df.columns:
        raise KeyError("Input data must contain 'revenue' and 'cost' columns.")
    
    # Calculate profit.
    df["profit"] = df["revenue"] - df["cost"]
    
    # Optionally calculate profit margin if requested.
    if args.calc_profit_margin:
        # Profit margin as a percentage; avoid division by zero.
        df["profit_margin"] = df.apply(
            lambda row: (row["profit"] / row["revenue"] * 100) if row["revenue"] != 0 else 0,
            axis=1
        )
        if args.verbose:
            print("Calculated profit margin.")
    
    # Create summary report using the specified aggregation method.
    agg_func = args.agg_method
    summary = pd.DataFrame({
        "Revenue": [df["revenue"].agg(agg_func)],
        "Cost": [df["cost"].agg(agg_func)],
        "Profit": [df["profit"].agg(agg_func)]
    })
    if args.calc_profit_margin:
        summary["Profit Margin"] = (
            summary["Profit"] / summary["Revenue"] * 100
            if summary["Revenue"].iloc[0] != 0 else 0
        )
    
    return df, summary

def generate_pivot(df, args):
    """
    If requested, generate a pivot table from the DataFrame.
    Uses --pivot-index and --pivot-values with the specified aggregation method.
    Returns the pivot table DataFrame (or None if not generated).
    """
    if args.generate_pivot:
        if args.pivot_index and args.pivot_values:
            pivot_index = args.pivot_index.lower()
            pivot_values = args.pivot_values.lower()
            if pivot_index in df.columns and pivot_values in df.columns:
                pivot_table = pd.pivot_table(
                    df,
                    index=pivot_index,
                    values=pivot_values,
                    aggfunc=args.agg_method
                )
                if args.verbose:
                    print("Generated pivot table.")
                return pivot_table.reset_index()
            else:
                print("Pivot table columns not found in data.")
    return None

def generate_monthly_totals(df, date_column, agg_method, calc_profit_margin, verbose=False):
    """
    Groups the DataFrame by month (using the specified date column) and computes totals.
    Also appends a final total row with overall sums.
    Returns a DataFrame with monthly totals and the final total row.
    """
    date_column = date_column.lower()
    if date_column not in df.columns:
        if verbose:
            print(f"Date column '{date_column}' not found; monthly totals will not be generated.")
        return None
    
    # Ensure the date column is datetime.
    df[date_column] = pd.to_datetime(df[date_column], errors="coerce")
    
    # Create a new column for month (e.g., "2023-01").
    df["month"] = df[date_column].dt.to_period('M').astype(str)
    
    # Group by month and sum revenue, cost, and profit.
    monthly_totals = df.groupby("month").agg({
        "revenue": "sum",
        "cost": "sum",
        "profit": "sum"
    }).reset_index()
    
    if calc_profit_margin:
        monthly_totals["profit_margin"] = monthly_totals.apply(
            lambda row: (row["profit"] / row["revenue"] * 100) if row["revenue"] != 0 else 0,
            axis=1
        )
    
    # Create a final total row.
    final_totals = {
        "month": "Final Total",
        "revenue": df["revenue"].sum(),
        "cost": df["cost"].sum(),
        "profit": df["profit"].sum()
    }
    if calc_profit_margin:
        final_totals["profit_margin"] = (final_totals["profit"] / final_totals["revenue"] * 100) if final_totals["revenue"] != 0 else 0
    
    final_totals_df = pd.DataFrame([final_totals])
    monthly_totals = pd.concat([monthly_totals, final_totals_df], ignore_index=True)
    
    if verbose:
        print("Generated monthly totals and final total row.")
    return monthly_totals

def save_report(cleaned_df, summary, pivot_df, monthly_totals, output_file, output_format, verbose=False, color_code=False):
    """
    Saves the cleaned data, summary, pivot table, and monthly totals.
    For Excel output, all are saved as separate sheets.
    For CSV output, each DataFrame is saved as a separate file.
    If color_code is True and output_format is 'xlsx', applies conditional formatting
    to the 'profit' column in the 'Cleaned Data' sheet.
    """
    if output_format == "xlsx":
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            cleaned_df.to_excel(writer, sheet_name="Cleaned Data", index=False)
            summary.to_excel(writer, sheet_name="Summary Report", index=False)
            if pivot_df is not None:
                pivot_df.to_excel(writer, sheet_name="Pivot Table", index=False)
            if monthly_totals is not None:
                monthly_totals.to_excel(writer, sheet_name="Monthly Totals", index=False)
        if verbose:
            print(f"Final report saved as '{output_file}'")
        
        # Apply color coding if requested.
        if color_code:
            try:
                from openpyxl import load_workbook
                from openpyxl.formatting.rule import CellIsRule
                from openpyxl.styles import PatternFill
                
                wb = load_workbook(output_file)
                ws = wb["Cleaned Data"]
                profit_col = None
                # Find the column letter for the "profit" header.
                for cell in ws[1]:
                    if cell.value and str(cell.value).strip().lower() == "profit":
                        profit_col = cell.column_letter
                        break
                
                if profit_col:
                    max_row = ws.max_row
                    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    
                    ws.conditional_formatting.add(f"{profit_col}2:{profit_col}{max_row}",
                        CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
                    ws.conditional_formatting.add(f"{profit_col}2:{profit_col}{max_row}",
                        CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
                    ws.conditional_formatting.add(f"{profit_col}2:{profit_col}{max_row}",
                        CellIsRule(operator='equal', formula=['0'], fill=yellow_fill))
                    
                    if verbose:
                        print("Applied color coding to 'profit' column in 'Cleaned Data' sheet.")
                else:
                    if verbose:
                        print("No 'profit' column found for color coding.")
                wb.save(output_file)
            except Exception as e:
                print(f"Error applying color coding: {e}")
                
    elif output_format == "csv":
        base, _ = os.path.splitext(output_file)
        cleaned_file = base + "_cleaned.csv"
        summary_file = base + "_summary.csv"
        cleaned_df.to_csv(cleaned_file, index=False)
        summary.to_csv(summary_file, index=False)
        if verbose:
            print(f"CSV reports saved as '{cleaned_file}' and '{summary_file}'.")
        if pivot_df is not None:
            pivot_file = base + "_pivot.csv"
            pivot_df.to_csv(pivot_file, index=False)
            if verbose:
                print(f"CSV pivot report saved as '{pivot_file}'.")
        if monthly_totals is not None:
            monthly_file = base + "_monthly_totals.csv"
            monthly_totals.to_csv(monthly_file, index=False)
            if verbose:
                print(f"CSV monthly totals saved as '{monthly_file}'.")
    else:
        raise ValueError("Unsupported output format.")

def archive_files(files, archive_folder, verbose=False):
    """
    Moves all files in the list to the archive folder.
    Creates the archive folder if it does not exist.
    """
    if not os.path.exists(archive_folder):
        os.makedirs(archive_folder)
    for file in files:
        base = os.path.basename(file)
        destination = os.path.join(archive_folder, base)
        os.rename(file, destination)
        if verbose:
            print(f"Archived '{file}' to '{destination}'.")

def main():
    parser = argparse.ArgumentParser(description="Advanced Excel Automation: Merge, Process, and Report")
    
    # Merging-related arguments.
    parser.add_argument("--data-folder", default="data",
                        help="Folder containing Excel files to merge (default: 'data')")
    parser.add_argument("--merged-file", default="merged_sales_data.xlsx",
                        help="Filename for the merged Excel file (default: 'merged_sales_data.xlsx')")
    parser.add_argument("--sheet-name", default=None,
                        help="Sheet name to read from each Excel file (default: first sheet if not specified)")
    
    # Output options.
    parser.add_argument("--output", default="final_report.xlsx",
                        help="Filename for the final processed report (default: 'final_report.xlsx')")
    parser.add_argument("--output-format", choices=["xlsx", "csv"], default="xlsx",
                        help="Output format for the report (default: xlsx)")
    
    # Data cleaning options.
    parser.add_argument("--dropna", action="store_true",
                        help="Drop rows with any missing data")
    parser.add_argument("--fillna", type=float, default=None,
                        help="Fill missing values with the specified value")
    
    # Date filtering options.
    parser.add_argument("--start-date", help="Start date for filtering (YYYY-MM-DD)")
    parser.add_argument("--end-date", help="End date for filtering (YYYY-MM-DD)")
    parser.add_argument("--date-column", default="date",
                        help="Column name for date filtering (default: 'date')")
    
    # Custom filtering: multiple --filter entries in the format "column:value"
    parser.add_argument("--filter", action="append",
                        help="Custom filter in the format column:value; can be specified multiple times")
    
    # Aggregation and calculated metrics.
    parser.add_argument("--agg-method", choices=["sum", "mean"], default="sum",
                        help="Aggregation method for summary (default: sum)")
    parser.add_argument("--calc-profit-margin", action="store_true",
                        help="Calculate profit margin (profit/revenue * 100)")
    
    # Pivot table options.
    parser.add_argument("--generate-pivot", action="store_true",
                        help="Generate a pivot table")
    parser.add_argument("--pivot-index", help="Column name to use as pivot table index (e.g., category)")
    parser.add_argument("--pivot-values", help="Column name to aggregate in pivot table (e.g., revenue)")
    
    # Logging and debugging options.
    parser.add_argument("--verbose", action="store_true",
                        help="Enable verbose output")
    parser.add_argument("--log-file", help="Log file to record detailed processing information")
    
    # File archiving.
    parser.add_argument("--archive", action="store_true",
                        help="Archive input files after processing")
    parser.add_argument("--archive-folder", default="archive",
                        help="Folder to move archived files (default: 'archive')")
    
    # User-defined transformations.
    parser.add_argument("--config", help="JSON configuration file for custom transformations")
    
    # Notification placeholder (email notification not implemented).
    parser.add_argument("--email", help="Email address to notify in case of errors (placeholder)")
    
    # New flag for enabling color coding.
    parser.add_argument("--color-code", action="store_true",
                        help="Apply color coding to the 'profit' column in the final Excel report")
    
    args = parser.parse_args()
    
    # Setup logging if requested.
    if args.log_file:
        logging.basicConfig(filename=args.log_file, level=logging.INFO,
                            format="%(asctime)s:%(levelname)s:%(message)s")
    
    if args.verbose:
        print("Starting Excel automation process with the following parameters:")
        print(args)
    
    try:
        # Merge Excel files from the specified folder.
        merged_df, files = merge_excel_files(args.data_folder, args.merged_file, args.sheet_name, args.verbose)
        
        # Process the merged data.
        cleaned_df, summary = process_data(merged_df, args)
        
        # Generate pivot table if requested.
        pivot_df = generate_pivot(cleaned_df, args)
        
        # Always generate monthly totals if the date column is present.
        monthly_totals = generate_monthly_totals(cleaned_df, args.date_column, args.agg_method, args.calc_profit_margin, args.verbose)
        
        # Save the final report (including monthly totals and color coding if requested).
        save_report(cleaned_df, summary, pivot_df, monthly_totals, args.output, args.output_format, args.verbose, args.color_code)
        
        # Archive original files if requested.
        if args.archive:
            archive_files(files, args.archive_folder, args.verbose)
        
        if args.verbose:
            print("Excel automation process completed successfully.")
    except Exception as e:
        error_message = f"An error occurred: {e}"
        print(error_message)
        if args.log_file:
            logging.error(error_message)
        if args.email:
            print(f"Notification sent to {args.email} (placeholder).")
        raise

if __name__ == "__main__":
    main()
